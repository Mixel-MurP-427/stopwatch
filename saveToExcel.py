#!/usr/bin/env python3
"""
saveToExcel.py

Read lists from a JSON file (default: `savaData.json`) and write them into
columns D, E, F, ... of a pre-existing Excel workbook.

Usage:
	python saveToExcel.py path/to/workbook.xlsx [--json path/to/savaData.json] [--sheet SHEETNAME] [--start-row N]

Dependencies:
	pip install openpyxl

Behavior:
	- JSON keys should be numeric strings like "1", "2", ...
	- Key "1" -> column D, "2" -> column E, etc.
	- Values are written top-down starting at --start-row (default 1).
"""

from __future__ import annotations

import argparse
import json
import sys
from typing import Any, Dict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def load_json(path: str) -> Dict[str, Any]:
	with open(path, "r", encoding="utf-8") as f:
		return json.load(f)


def write_columns_from_data(ws, data: Dict[str, Any], start_row: int = 1) -> int:
	"""Write lists found in `data` to worksheet `ws`.

	Returns number of lists written.
	"""
	# Sort keys numerically if possible
	keys = sorted(data.keys(), key=lambda k: int(k) if str(k).isdigit() else float("inf"))

	written = 0
	for key in keys:
		try:
			idx = int(key)
		except Exception:
			print(f"Skipping non-integer key: {key}", file=sys.stderr)
			continue

		col_index = idx + 3  # list 1 -> col 4 (D), list 2 -> col 5 (E), ...
		col_letter = get_column_letter(col_index)

		values = data.get(key)
		if not isinstance(values, list):
			print(f"Key {key} does not contain a list; skipping.", file=sys.stderr)
			continue

		for i, val in enumerate(values):
			row = start_row + i
			ws[f"{col_letter}{row}"] = val

		written += 1

	return written


def main() -> None:
	parser = argparse.ArgumentParser(description="Write JSON lists into columns D/E/... of an existing xlsx file")
	parser.add_argument("xlsx", help="Path to existing .xlsx workbook to update")
	parser.add_argument("--json", default="savaData.json", help="Path to JSON file (default: savaData.json)")
	parser.add_argument("--sheet", default=None, help="Worksheet name to write to (default: active sheet)")
	parser.add_argument("--start-row", type=int, default=2, help="Row to start writing at (default: 2)")

	args = parser.parse_args()

	try:
		data = load_json(args.json)
	except Exception as exc:
		print(f"Failed to load JSON file '{args.json}': {exc}", file=sys.stderr)
		sys.exit(2)

	try:
		wb = load_workbook(args.xlsx)
	except Exception as exc:
		print(f"Failed to open workbook '{args.xlsx}': {exc}", file=sys.stderr)
		sys.exit(3)

	ws = wb[args.sheet] if args.sheet and args.sheet in wb.sheetnames else wb.active

	written = write_columns_from_data(ws, data, start_row=args.start_row)

	try:
		wb.save(args.xlsx)
	except Exception as exc:
		print(f"Failed to save workbook '{args.xlsx}': {exc}", file=sys.stderr)
		sys.exit(4)

	print(f"Wrote {written} list(s) to '{args.xlsx}' on sheet '{ws.title}', starting at row {args.start_row}.")


if __name__ == "__main__":
	main()

